import pandas as pd
import numpy as np
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .db import get_engine


COLUMN_WIDTHS = {
    'A': 8,      # Internal Running No
    'B': 20,     # Sales Order No
    'C': 20,     # External ID
	'D': 18,
	'E': 30,
	'F': 18,
	'G': 30,
	'H': 18,
	'I': 30,
    'AD': 18,
    'AE': 30,
    'AF': 10,
    'AG':18,
    'AH':10,
    'AI':18,
    'BA': 25,
    'BC': 20
}


def parse_sap_number(val):
    """Parse SAP number format e.g. '  0.560000000-' -> -0.56"""
    if pd.isna(val):
        return np.nan
    s = str(val).strip()
    if s == '':
        return np.nan
    if s.endswith('-'):
        s = '-' + s[:-1]
    try:
        return float(s)
    except ValueError:
        return np.nan


OUTPUT_COLUMNS = [
    'Product',                                          # A
    'Internal Running No',                              # B
    'External ID of Sales Order',                       # C
    'Internal ID of Customer',                          # D
    'Customer',                                         # E
    'Internal ID of Ship to Customer',                  # F
    'Ship to Customer',                                 # G
    'Internal ID of Shipping Label New',                # H
    'Shipping Label',                                   # I
    'Date',                                             # J
    'Domestic/Export',                                   # K
    'Internal ID of Subsidiary',                        # L
    'Subsidiary',                                       # M
    'Internal ID Payment Terms',                        # N
    'Payment Terms',                                    # O
    'Internal ID IncotermsTerms',                       # P
    'IncotermsTerms',                                   # Q
    'Internal ID Ship from Country',                    # R
    'Ship from Country',                                # S
    'Internal ID Ship to Country',                      # T
    'Ship to Country',                                  # U
    'Internal ID Port',                                 # V
    'Port',                                             # W
    'ETD Date',                                         # X
    'Currency',                                         # Y
    'Exchange Rate',                                    # Z
    'Memo',                                             # AA
    'PO Number',                                        # AB
    'PO Date',                                          # AC
    'Internal ID of Sales Ref.',                        # AD
    'Sales Ref.',                                       # AE
    'Internal ID of Sales Channel',                     # AF
    'Sales Channel',                                    # AG
    'Internal ID of Distribution Channel',              # AH
    'Distribution Channel',                             # AI
    'Internal ID of Customer Country',                  # AJ
    'Customer Country',                                 # AK
    'Internal ID of Continent',                         # AL
    'Continent',                                        # AM
    'Internal ID of Customer Region',                   # AN
    'Customer Region',                                  # AO
    'Internal ID of Department',                        # AP
    'Department',                                       # AQ
    'Internal ID of Location (Branch)',                 # AR
    'Location (Branch)',                                # AS
    'Internal ID Cluster CEO',                          # AT
    'Cluster CEO',                                      # AU
    'Remark to Warehouse',                              # AV
    'Remark to Transport',                              # AW
    'Remark to Account',                                # AX
    'Place to',                                         # AY
    'Internal ID of Item',                              # AZ
    'Item',                                             # BA
    'Internal ID of Location (Line)',                   # BB
    'Location (Line)',                                  # BC
    'Quantity',                                         # BD
    'Unit',                                             # BE
    'Internal ID ofSPS Sale Discount Reference',        # BF
    'SPS Sale Discount Reference',                      # BG
    'SPS STD Sales Unit Price',                         # BH
    'SPS STD Sales Total',                              # BI
    'Unit Price after Item Discount (Exc.VAT)',         # BJ
    'Amount',                                           # BK
    'Tax Code 1',                                       # BL
    'Tax Amt',                                          # BM
    'Gross Amount',                                     # BN
    'Amount/TON',                                       # BO
    'Conv.of Unit (EA/PC)',                             # BP
    'Conv. of Unit (Pack)',                             # BQ
    'Conv. of Unit (KG)',                               # BR
    'Conv.of Unit (Ream)',                              # BS
    'Conv.of Unit (Box)',                               # BT
    'Conv.of Unit (TON)',                               # BU
    'Conv.of Unit (PAL)',                               # BV
    'Net Weight',                                       # BW
    'Gross Weight',                                     # BX
    'SPS Sales Discount Total',                         # BY
    'ID of Discount/Markup 1',                          # BZ
    'Code of Discount/Markup 1',                        # CA
    'SPS Net Discount/Markup 1',                        # CB
    'ID of Discount/Markup 2',                          # CC
    'Code of Discount/Markup 2',                        # CD
    'SPS Net Discount/Markup 2',                        # CE
    'ID of Discount/Markup 3',                          # CF
    'Code of Discount/Markup 3',                        # CG
    'SPS Net Discount/Markup 3',                        # CH
    'ID of Discount/Markup 4',                          # CI
    'Code of Discount/Markup 4',                        # CJ
    'SPS Net Discount/Markup 4',                        # CK
    'ID of Discount/Markup 5',                          # CL
    'Code of Discount/Markup 5',                        # CM
    'SPS Net Discount/Markup 5',                        # CN
    'ID of Discount/Markup 6',                          # CO
    'Code of Discount/Markup 6',                        # CP
    'SPS Net Discount/Markup 6',                        # CQ
    'Internal ID of Promotion Order Code1',             # CR
    'Promotion Order Code1',                            # CS
    'Promotion Order Amount 1',                         # CT
    'Internal ID of Promotion Order Code2',             # CU
    'Promotion Order Code2',                            # CV
    'Promotion Order Amount 2',                         # CW
    'Internal ID of Promotion Order Code3',             # CX
    'Promotion Order Code3',                            # CY
    'Promotion Order Amount 3',                         # CZ
    'Promotion Order Amount Total',                     # DA
    'Internal ID of Promotion Code (FOC)',              # DB
    'Promotion Code (FOC)',                             # DC
    'Blank 2',                                          # DD
    'Tax Code 2',                                       # DE
]

# SAP report column indices
SAP_COL_SALES_ORDER_NO       = 0
SAP_COL_MATERIAL_NO          = 2
SAP_COL_DIST_CHANNEL_NAME    = 6
SAP_COL_PAYMENT_TERM_CODE    = 7
SAP_COL_SOLD_TO_NO           = 9
SAP_COL_SHIP_TO_NO           = 11
SAP_COL_SHIPPING_LABEL       = 13
SAP_COL_SALES_EMPLOYEE_NAME  = 15
SAP_COL_PO_NUMBER            = 16
SAP_COL_PLANT                = 17
SAP_COL_SO_DATE              = 20
SAP_COL_PO_DATE              = 21
SAP_COL_REQUEST_DELIVERY_DATE = 22
SAP_COL_ORDER_QUANTITY       = 23
SAP_COL_PENDING_QUANTITY     = 24
SAP_COL_SALES_UNIT           = 25
SAP_COL_AMOUNT               = 35
SAP_COL_ZP01                 = 41
SAP_COL_ZP01_AMOUNT          = 43
SAP_COL_ZP01_VALUE           = 47
SAP_COL_ZC01                 = 49
SAP_COL_ZC01_AMOUNT          = 51
SAP_COL_ZC01_VALUE           = 55
SAP_COL_ZD13                 = 57
SAP_COL_ZD13_VALUE           = 63


def _recalc_running_no(output_df):
    running_nos = []
    current_so = None
    counter = 0
    for ext_id in output_df['External ID of Sales Order'].values:
        if ext_id == current_so:
            counter += 1
        else:
            current_so = ext_id
            counter = 1
        running_nos.append(counter)
    output_df['Internal Running No'] = (
        output_df['External ID of Sales Order'] + '_' + [str(n) for n in running_nos]
    )


def generate_opening_so(excel_path: str, output_path: str) -> dict:
    """Generate Opening SO Excel from SAP report.

    Returns a dict with 'status', 'message', 'rows' count.
    """
    engine = get_engine()

    # ===== 1. Read SAP report =====
    sap_df = pd.read_excel(excel_path, sheet_name='SAP report')
    col_name_so = sap_df.columns[SAP_COL_SALES_ORDER_NO]
    sap_df = sap_df.dropna(subset=[col_name_so])
    sap_df[col_name_so] = sap_df[col_name_so].astype(int).astype(str)
    sap_df = sap_df.sort_values(by=col_name_so).reset_index(drop=True)

    # ===== 2. Load lookup tables =====
    customer_df = pd.read_sql(
        "SELECT internal_id, old_code, code, customer_name FROM customer_master", con=engine
    )
    valid_customers = customer_df.dropna(subset=['old_code'])
    valid_customers = valid_customers[valid_customers['old_code'].astype(str).str.strip() != '']
    old_code_to_internal_id = dict(zip(valid_customers['old_code'].astype(str), valid_customers['internal_id']))
    old_code_to_code = dict(zip(valid_customers['old_code'].astype(str), valid_customers['code']))
    internal_id_to_customer_name = dict(zip(customer_df['internal_id'], customer_df['customer_name']))

    shipping_label_df = pd.read_sql(
        "SELECT address_internal_id, ship_to_code, address_label, ship_to_name FROM shipping_label_master", con=engine
    )
    valid_shipping = shipping_label_df.dropna(subset=['address_label'])
    valid_shipping = valid_shipping[valid_shipping['address_label'].astype(str).str.strip() != '']
    shipping_label_choices = valid_shipping['address_label'].astype(str).str.strip().tolist()
    shipping_label_records = valid_shipping[['address_internal_id']].to_dict('records')

    payment_term_df = pd.read_sql(
        "SELECT internal_id, payment_term_code, name FROM payment_term_master", con=engine
    )
    valid_payment = payment_term_df.dropna(subset=['payment_term_code'])
    valid_payment = valid_payment[valid_payment['payment_term_code'].astype(str).str.strip() != '']
    payment_code_to_internal_id = dict(zip(valid_payment['payment_term_code'].astype(str), valid_payment['internal_id']))
    payment_code_to_name = dict(zip(valid_payment['payment_term_code'].astype(str), valid_payment['name']))

    sales_rep_df = pd.read_sql(
        "SELECT internal_id, employee_id, name, first_name_th, last_name_th FROM sales_rep_master", con=engine
    )
    sales_rep_df['full_name_th'] = (
        sales_rep_df['first_name_th'].fillna('') + sales_rep_df['last_name_th'].fillna('')
    ).str.strip()
    valid_sales_rep = sales_rep_df[sales_rep_df['full_name_th'] != ''].copy()
    sales_rep_choices = valid_sales_rep['full_name_th'].tolist()
    sales_rep_records = valid_sales_rep[['internal_id', 'employee_id', 'name']].to_dict('records')

    item_df = pd.read_sql(
        "SELECT internal_id, item_code, item_name, old_item_code FROM item_master", con=engine
    )
    valid_items = item_df.dropna(subset=['old_item_code'])
    valid_items = valid_items[valid_items['old_item_code'].astype(str).str.strip() != '']
    old_item_code_to_internal_id = dict(zip(valid_items['old_item_code'].astype(str), valid_items['internal_id']))
    old_item_code_to_item_code = dict(zip(valid_items['old_item_code'].astype(str), valid_items['item_code']))

    location_df = pd.read_sql("SELECT plant, location_id, location_name FROM location_master", con=engine)
    plant_to_location_id = dict(zip(location_df['plant'].astype(str), location_df['location_id']))
    plant_to_location_name = dict(zip(location_df['plant'].astype(str), location_df['location_name']))

    try:
        sales_dist_df = pd.read_sql(
            "SELECT internal_id, distribution_channel, sale_channel_id, sale_channel FROM sales_dist_master",
            con=engine,
        )
        sales_dist_df['distribution_channel'] = sales_dist_df['distribution_channel'].astype(str).str.strip()
        sales_dist_unique = sales_dist_df.drop_duplicates(subset=['distribution_channel'], keep='first')
        dc_to_sc_id = dict(zip(sales_dist_unique['distribution_channel'], sales_dist_unique['sale_channel_id']))
        dc_to_sc_name = dict(zip(sales_dist_unique['distribution_channel'], sales_dist_unique['sale_channel']))
        dc_to_internal_id = dict(zip(sales_dist_unique['distribution_channel'], sales_dist_unique['internal_id']))
    except Exception:
        dc_to_sc_id = {}
        dc_to_sc_name = {}
        dc_to_internal_id = {}

    # Dynamic discount mappings from discount_master (detected during /api/validate)
    try:
        discount_master_df = pd.read_sql(
            "SELECT sap_discount, internal_id, discount_name, discount_type, column_index FROM discount_master",
            con=engine,
        )
    except Exception as e:
        raise ValueError("Cannot load discount_master. Please run header validation before generate.") from e

    discount_master_df = discount_master_df.copy()
    discount_master_df['sap_discount'] = discount_master_df['sap_discount'].astype(str).str.strip().str.upper()
    discount_master_df['internal_id'] = discount_master_df['internal_id'].fillna('').astype(str).str.strip()
    discount_master_df['discount_name'] = discount_master_df['discount_name'].fillna('').astype(str).str.strip()
    discount_master_df['discount_type'] = discount_master_df['discount_type'].astype(str).str.strip().str.lower()
    discount_master_df['column_index'] = pd.to_numeric(discount_master_df['column_index'], errors='coerce')
    discount_master_df = discount_master_df.dropna(subset=['sap_discount', 'discount_type', 'column_index'])
    discount_master_df = discount_master_df[discount_master_df['sap_discount'] != '']
    discount_master_df['column_index'] = discount_master_df['column_index'].astype(int)
    discount_master_df = discount_master_df.sort_values(by='column_index')

    item_discounts = discount_master_df[discount_master_df['discount_type'] == 'by item'][
        ['sap_discount', 'column_index']
    ].to_dict('records')
    order_discounts = discount_master_df[discount_master_df['discount_type'] == 'by order'][
        ['sap_discount', 'column_index', 'internal_id', 'discount_name']
    ].to_dict('records')

    if len(item_discounts) > 6:
        raise ValueError(
            f"Found {len(item_discounts)} item discounts in discount_master, but output supports only 6."
        )
    if len(order_discounts) > 3:
        raise ValueError(
            f"Found {len(order_discounts)} order discounts in discount_master, but output supports only 3."
        )

    sap_col_count = sap_df.shape[1]

    def _require_sap_col(idx: int, label: str) -> int:
        if idx < 0 or idx >= sap_col_count:
            raise ValueError(
                f"SAP column index out of range for {label}: {idx}. "
                f"Available SAP columns: 0-{sap_col_count - 1}"
            )
        return idx

    # ===== 3. Build output DataFrame =====
    output_df = pd.DataFrame(columns=OUTPUT_COLUMNS)

    # (C) External ID of Sales Order
    output_df['External ID of Sales Order'] = 'SO_DA_' + sap_df.iloc[:, SAP_COL_SALES_ORDER_NO].values

    # (A) Product
    output_df['Product'] = 'CZ'

    # (B) Internal Running No
    _recalc_running_no(output_df)

    # (D)(E) Customer
    sold_to_nos = sap_df.iloc[:, SAP_COL_SOLD_TO_NO].values
    ids_cust, names_cust = [], []
    for sold_to in sold_to_nos:
        if pd.isna(sold_to) or str(sold_to).strip() == '':
            ids_cust.append(None); names_cust.append(None)
        else:
            s = str(int(sold_to)) if isinstance(sold_to, float) else str(sold_to)
            iid = old_code_to_internal_id.get(s)
            ids_cust.append(iid)
            names_cust.append(internal_id_to_customer_name.get(iid) if iid else None)
    output_df['Internal ID of Customer'] = ids_cust
    output_df['Customer'] = names_cust

    # (F)(G) Ship to Customer
    ship_to_nos = sap_df.iloc[:, SAP_COL_SHIP_TO_NO].values
    ids_ship, names_ship = [], []
    for ship_to in ship_to_nos:
        if pd.isna(ship_to) or str(ship_to).strip() == '':
            ids_ship.append(None); names_ship.append(None)
        else:
            s = str(int(ship_to)) if isinstance(ship_to, float) else str(ship_to)
            iid = old_code_to_internal_id.get(s)
            ids_ship.append(iid)
            names_ship.append(internal_id_to_customer_name.get(iid) if iid else None)
    output_df['Internal ID of Ship to Customer'] = ids_ship
    output_df['Ship to Customer'] = names_ship

    # (H)(I) Shipping Label
    sap_shipping_labels = sap_df.iloc[:, SAP_COL_SHIPPING_LABEL].values
    label_ids = []
    for label in sap_shipping_labels:
        if pd.isna(label) or str(label).strip() == '':
            label_ids.append(None)
        else:
            result = process.extractOne(str(label).strip(), shipping_label_choices, scorer=fuzz.ratio)
            if result and result[1] >= 80:
                idx = shipping_label_choices.index(result[0])
                label_ids.append(shipping_label_records[idx]['address_internal_id'])
            else:
                label_ids.append(None)
    output_df['Internal ID of Shipping Label New'] = label_ids
    output_df['Shipping Label'] = sap_shipping_labels

    # (J) Date
    so_dates = pd.to_datetime(sap_df.iloc[:, SAP_COL_SO_DATE], errors='coerce')
    output_df['Date'] = so_dates.dt.strftime('%d/%m/%y').values

    # (K)-(M) Constants
    output_df['Domestic/Export'] = 'Domestic'
    output_df['Internal ID of Subsidiary'] = 2
    output_df['Subsidiary'] = 'Double A (1991) PLC'

    # (N)(O) Payment Terms
    pt_codes = sap_df.iloc[:, SAP_COL_PAYMENT_TERM_CODE].values
    pt_ids, pt_names = [], []
    for pt in pt_codes:
        if pd.isna(pt) or str(pt).strip() == '':
            pt_ids.append(None); pt_names.append(None)
        else:
            s = str(pt).strip()
            pt_ids.append(payment_code_to_internal_id.get(s))
            pt_names.append(payment_code_to_name.get(s))
    output_df['Internal ID Payment Terms'] = pt_ids
    output_df['Payment Terms'] = pt_names

    # (P)(Q) Incoterms
    output_df['Internal ID IncotermsTerms'] = 5
    output_df['IncotermsTerms'] = 'DAP'

    # (S) Ship from Country
    output_df['Ship from Country'] = 'Thailand'

    # (X) ETD Date
    etd_dates = pd.to_datetime(sap_df.iloc[:, SAP_COL_REQUEST_DELIVERY_DATE], errors='coerce')
    output_df['ETD Date'] = etd_dates.dt.strftime('%d/%m/%y').values

    # (Y)(Z) Currency
    output_df['Currency'] = 'THB'
    output_df['Exchange Rate'] = 1

    # (AB)(AC) PO
    output_df['PO Number'] = sap_df.iloc[:, SAP_COL_PO_NUMBER].values
    po_dates = pd.to_datetime(sap_df.iloc[:, SAP_COL_PO_DATE], errors='coerce')
    output_df['PO Date'] = po_dates.dt.strftime('%d/%m/%y').values

    # (AD)(AE) Sales Ref - fuzzy match
    emp_names = sap_df.iloc[:, SAP_COL_SALES_EMPLOYEE_NAME].values
    sr_ids, sr_names = [], []
    for emp in emp_names:
        if pd.isna(emp) or str(emp).strip() == '':
            sr_ids.append(None); sr_names.append(None)
        else:
            result = process.extractOne(str(emp).strip(), sales_rep_choices, scorer=fuzz.ratio)
            if result and result[1] >= 80:
                idx = sales_rep_choices.index(result[0])
                rec = sales_rep_records[idx]
                sr_ids.append(rec['internal_id'])
                sr_names.append(f"{rec['employee_id']} {rec['name']}")
            else:
                sr_ids.append(None); sr_names.append(None)
    output_df['Internal ID of Sales Ref.'] = sr_ids
    output_df['Sales Ref.'] = sr_names

    # (AF)(AG)(AH)(AI) Sales Channel / Distribution Channel
    dist_channel_names = sap_df.iloc[:, SAP_COL_DIST_CHANNEL_NAME].values
    af_ids, ag_names, ah_ids, ai_names = [], [], [], []
    for dcn in dist_channel_names:
        if pd.isna(dcn) or str(dcn).strip() == '':
            af_ids.append(None); ag_names.append(None)
            ah_ids.append(None); ai_names.append(None)
        else:
            s = str(dcn).strip()
            af_ids.append(dc_to_sc_id.get(s))
            ag_names.append(dc_to_sc_name.get(s))
            ah_ids.append(dc_to_internal_id.get(s))
            ai_names.append(s if dc_to_internal_id.get(s) is not None else None)
    output_df['Internal ID of Sales Channel'] = af_ids
    output_df['Sales Channel'] = ag_names
    output_df['Internal ID of Distribution Channel'] = ah_ids
    output_df['Distribution Channel'] = ai_names

    # (AJ)-(AM) Country / Continent constants
    output_df['Internal ID of Customer Country'] = 324
    output_df['Customer Country'] = 'Thailand'
    output_df['Internal ID of Continent'] = 2
    output_df['Continent'] = 'Asia'

    # (AZ)(BA) Item
    mat_nos = sap_df.iloc[:, SAP_COL_MATERIAL_NO].values
    item_ids, item_codes = [], []
    for m in mat_nos:
        if pd.isna(m) or str(m).strip() == '':
            item_ids.append(None); item_codes.append(None)
        else:
            s = str(int(m)) if isinstance(m, float) else str(m)
            item_ids.append(old_item_code_to_internal_id.get(s))
            item_codes.append(old_item_code_to_item_code.get(s))
    output_df['Internal ID of Item'] = item_ids
    output_df['Item'] = item_codes

    # (BB)(BC) Location
    plants = sap_df.iloc[:, SAP_COL_PLANT].values
    loc_ids, loc_names = [], []
    for p in plants:
        if pd.isna(p) or str(p).strip() == '':
            loc_ids.append(None); loc_names.append(None)
        else:
            s = str(int(p)) if isinstance(p, float) else str(p)
            loc_ids.append(plant_to_location_id.get(s))
            loc_names.append(plant_to_location_name.get(s))
    output_df['Internal ID of Location (Line)'] = loc_ids
    output_df['Location (Line)'] = loc_names

    # (BD)(BE) Quantity / Unit
    output_df['Quantity'] = sap_df.iloc[:, SAP_COL_PENDING_QUANTITY].values
    output_df['Unit'] = sap_df.iloc[:, SAP_COL_SALES_UNIT].values

    # (BH) SPS STD Sales Unit Price
    output_df['SPS STD Sales Unit Price'] = sap_df.iloc[:, SAP_COL_AMOUNT].apply(parse_sap_number)

    # (BI) SPS STD Sales Total
    output_df['SPS STD Sales Total'] = output_df['Quantity'] * output_df['SPS STD Sales Unit Price']

    # (BJ), (CA-CQ) Dynamic by-item discount mapping from discount_master
    # Phase 1: parse all item discounts from SAP columns
    item_discount_parsed = []
    item_discount_amount_sum = pd.Series(0.0, index=sap_df.index)
    for discount in item_discounts:
        sap_code = discount['sap_discount']
        base_idx = _require_sap_col(int(discount['column_index']), f"{sap_code} base")
        amount_idx = _require_sap_col(base_idx + 2, f"{sap_code} amount")
        value_idx = _require_sap_col(base_idx + 6, f"{sap_code} value")

        sap_code_col = sap_df.iloc[:, base_idx].astype(str).str.strip().str.upper()
        has_discount = sap_code_col.str.contains(sap_code, na=False)
        net_values = sap_df.iloc[:, value_idx].apply(parse_sap_number)
        amount_values = sap_df.iloc[:, amount_idx].apply(parse_sap_number).fillna(0)
        item_discount_amount_sum = item_discount_amount_sum + amount_values.where(has_discount, 0)

        item_discount_parsed.append({
            'sap_code': sap_code,
            'has': has_discount,
            'net': net_values,
        })

    # Phase 2: per-row compact into slots 1..6
    for row_idx in range(len(output_df)):
        slot = 1
        for d in item_discount_parsed:
            if d['has'].iloc[row_idx]:
                output_df.at[row_idx, f'Code of Discount/Markup {slot}'] = d['sap_code']
                output_df.at[row_idx, f'SPS Net Discount/Markup {slot}'] = d['net'].iloc[row_idx]
                slot += 1

    output_df['Unit Price after Item Discount (Exc.VAT)'] = (
        output_df['SPS STD Sales Unit Price'] + item_discount_amount_sum
    )

    # (BK) Amount
    output_df['Amount'] = output_df['Unit Price after Item Discount (Exc.VAT)'] * output_df['Quantity']

    # (BL)(BM)(BN) Tax
    output_df['Tax Code 1'] = 'VAT_TH:VAT 7%'
    output_df['Tax Amt'] = output_df['Amount'] * 0.07
    output_df['Gross Amount'] = output_df['Amount'] + output_df['Tax Amt']

    # (BY) SPS Sales Discount Total
    discount_cols = [
        'SPS Net Discount/Markup 1', 'SPS Net Discount/Markup 2',
        'SPS Net Discount/Markup 3', 'SPS Net Discount/Markup 4',
        'SPS Net Discount/Markup 5', 'SPS Net Discount/Markup 6',
    ]
    output_df['SPS Sales Discount Total'] = output_df[discount_cols].fillna(0).sum(axis=1)

    # (CT-CZ) Dynamic by-order discount mapping from discount_master
    # Phase 1: parse all order discounts from SAP columns
    order_discount_parsed = []
    for discount in order_discounts:
        sap_code = discount['sap_discount']
        base_idx = _require_sap_col(int(discount['column_index']), f"{sap_code} base")
        value_idx = _require_sap_col(base_idx + 6, f"{sap_code} value")

        sap_code_col = sap_df.iloc[:, base_idx].astype(str).str.strip().str.upper()
        has_discount = sap_code_col.str.contains(sap_code, na=False)
        promo_values = sap_df.iloc[:, value_idx].apply(parse_sap_number)

        order_discount_parsed.append({
            'sap_code': sap_code,
            'has': has_discount,
            'value': promo_values,
        })

    # Phase 2: per-row compact into Promotion Order Code/Amount 1..3
    for row_idx in range(len(output_df)):
        slot = 1
        for d in order_discount_parsed:
            if d['has'].iloc[row_idx]:
                output_df.at[row_idx, f'Promotion Order Code{slot}'] = d['sap_code']
                output_df.at[row_idx, f'Promotion Order Amount {slot}'] = d['value'].iloc[row_idx]
                slot += 1

    # (DA) Promotion Order Amount Total
    promo_cols = ['Promotion Order Amount 1', 'Promotion Order Amount 2', 'Promotion Order Amount 3']
    output_df['Promotion Order Amount Total'] = output_df[promo_cols].fillna(0).sum(axis=1)

    # ===== 3.5 Insert subtotal rows =====
    has_promo = output_df[promo_cols].notna().any(axis=1) & (output_df[promo_cols].fillna(0) != 0).any(axis=1)
    copy_cols = (
        [OUTPUT_COLUMNS[0]]
        + OUTPUT_COLUMNS[2:7]
        + OUTPUT_COLUMNS[8:51]
        + OUTPUT_COLUMNS[53:55]
    )
    sum_cols = [
        'SPS Sales Discount Total', 'SPS Net Discount/Markup 1', 'SPS Net Discount/Markup 2',
        'SPS Net Discount/Markup 3', 'SPS Net Discount/Markup 4', 'SPS Net Discount/Markup 5',
        'SPS Net Discount/Markup 6', 'Promotion Order Amount 1', 'Promotion Order Amount 2',
        'Promotion Order Amount 3', 'Promotion Order Amount Total',
    ]

    inserted_row_indices = []
    promotion_row_indices = []

    if has_promo.any():
        so_with_promo = set(output_df.loc[has_promo, 'External ID of Sales Order'].unique())
        so_sums = output_df[output_df['External ID of Sales Order'].isin(so_with_promo)] \
            .groupby('External ID of Sales Order')[sum_cols].sum()

        new_rows = []
        is_inserted = []
        prev_so = None

        def _build_subtotal(last_row, so_ext_id):
            dup = {col: None for col in OUTPUT_COLUMNS}
            for col in copy_cols:
                dup[col] = last_row[col]
            dup['Internal ID of Item'] = -2
            dup['Item'] = 'Subtotal'
            for col in sum_cols:
                dup[col] = so_sums.loc[so_ext_id, col]
            return dup

        for idx in range(len(output_df)):
            current_so = output_df.iloc[idx]['External ID of Sales Order']
            if prev_so is not None and current_so != prev_so and prev_so in so_with_promo:
                new_rows.append(_build_subtotal(new_rows[-1], prev_so))
                is_inserted.append(True)
            new_rows.append(output_df.iloc[idx].to_dict())
            is_inserted.append(False)
            prev_so = current_so

        if prev_so in so_with_promo:
            new_rows.append(_build_subtotal(new_rows[-1], prev_so))
            is_inserted.append(True)

        output_df = pd.DataFrame(new_rows, columns=OUTPUT_COLUMNS).reset_index(drop=True)
        _recalc_running_no(output_df)
        inserted_row_indices = [i for i, f in enumerate(is_inserted) if f]

        # ===== 3.6 Insert promotion rows =====
        subtotal_set = set(inserted_row_indices)
        subtotal_with_promo = [
            i for i in inserted_row_indices
            if output_df.loc[i, promo_cols].notna().any()
            and (output_df.loc[i, promo_cols].fillna(0) != 0).any()
        ]

        if subtotal_with_promo:
            subtotal_with_promo_set = set(subtotal_with_promo)
            new_rows2 = []
            row_types = []

            for idx in range(len(output_df)):
                new_rows2.append(output_df.iloc[idx].to_dict())
                row_types.append('subtotal' if idx in subtotal_set else 'normal')

                if idx in subtotal_with_promo_set:
                    for promo_idx, promo_col in enumerate(promo_cols):
                        promo_val = output_df.at[idx, promo_col]
                        if pd.isna(promo_val) or promo_val == 0:
                            continue

                        promo_row = output_df.iloc[idx].to_dict()
                        order_discount = order_discounts[promo_idx] if promo_idx < len(order_discounts) else {}
                        promo_row['Internal ID of Item'] = order_discount.get('internal_id', '')
                        promo_row['Item'] = order_discount.get('discount_name', '')
                        promo_row['SPS STD Sales Total'] = promo_val
                        promo_row['Amount'] = promo_val
                        promo_row['Tax Code 1'] = 'VAT_TH:VAT 7%'
                        promo_row['Tax Amt'] = promo_val * 0.07
                        promo_row['Gross Amount'] = promo_val + promo_val * 0.07
                        new_rows2.append(promo_row)
                        row_types.append('promotion')

            output_df = pd.DataFrame(new_rows2, columns=OUTPUT_COLUMNS).reset_index(drop=True)
            _recalc_running_no(output_df)
            inserted_row_indices = [i for i, t in enumerate(row_types) if t == 'subtotal']
            promotion_row_indices = [i for i, t in enumerate(row_types) if t == 'promotion']

    # ===== 4. Write to Excel =====
    output_df.to_excel(output_path, index=False, sheet_name='Opening SO')

    wb = load_workbook(output_path)
    ws = wb['Opening SO']

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    if inserted_row_indices or promotion_row_indices:
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        pink_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        num_cols = len(OUTPUT_COLUMNS)
        for row_idx in inserted_row_indices:
            for col in range(1, num_cols + 1):
                ws.cell(row=row_idx + 2, column=col).fill = green_fill
        for row_idx in promotion_row_indices:
            for col in range(1, num_cols + 1):
                ws.cell(row=row_idx + 2, column=col).fill = pink_fill

    wb.save(output_path)

    return {
        "status": "success",
        "message": f"Generated {len(output_df)} rows",
        "rows": len(output_df),
        "subtotal_rows": len(inserted_row_indices),
        "promotion_rows": len(promotion_row_indices),
    }
